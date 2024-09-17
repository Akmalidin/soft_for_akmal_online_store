# LOGIN AND REGISTRATIONS IMPORT FILES 
from django.contrib.auth import authenticate, login
# DERICTIONS IMPORT FILES
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
# EXPORT FILES
from django.db.models import Q
import openpyxl, csv, io, chardet
import pandas as pd
from io import BytesIO
from openpyxl.styles import Border, Side
from django.contrib import messages
from collections import defaultdict
# MY MODELS
from .models import CustomUser, Product, Category, TypeOfPrice
from .forms import ImportProductsForm



@login_required
def index(request):
    query = request.GET.get('q', '')
    category_id = request.GET.get('category_id', '')
    sort_by = request.GET.get('sort_by', 'name')
    order_by = request.GET.get('order_by', 'asc')

    products = Product.objects.all()

    if query:
        search_terms = query.split()
        for term in search_terms:
            products = products.filter(
                Q(name__icontains=term) | Q(articul__icontains=term)
            )

    if category_id:
        category = get_object_or_404(Category, id=category_id)
        products = products.filter(category=category)

    # Применяем сортировку
    if sort_by:
        if sort_by == 'category':
            products = products.order_by('category__name' if order_by == 'asc' else '-category__name')
        elif sort_by == 'articul':
            products = products.order_by('articul' if order_by == 'asc' else '-articul')
        else:
            products = products.order_by(sort_by if order_by == 'asc' else '-' + sort_by)

    categories = Category.objects.all()
    context = {
        'products': products,
        'categories': categories,
        'query': query,
        'category_id': category_id,
        'order_by': order_by
    }
    return render(request, 'pages/index.html', context)

def user_login(request):
    users = CustomUser.objects.all()
    if request.method == 'POST':
        # Регистрация
        if request.POST:
            name = request.POST.get('name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            password = request.POST.get('password')
            profile_photo = request.FILES.get('profile_photo')
            address = request.POST.get('address')

            if email and phone_number and password:
                user = CustomUser.objects.create(
                    name=name,
                    last_name=last_name,
                    email=email,
                    phone_number=phone_number,
                    profile_photo=profile_photo,
                    address=address
                )
                user.set_password(password)  # Хешируем пароль
                user.save()
                return redirect(index)

        # Авторизация
        if request.POST.get('sign_in'):
            email = request.POST.get('email')
            password = request.POST.get('password')
            user = authenticate(request, username=email, password=password)

            if user is not None:
                login(request, user)
                return redirect(index)
            else:
                return render(request, 'login.html', {'error': 'Неверные учетные данные'})
    
    return render(request, 'login.html', locals())

@login_required
def export_products_to_excel(request):
    # Создание книги и листа
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Прайс лист'

    # Определение стиля границы
    thin_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # Заголовки столбцов
    columns = [
        'ID', 'Категория', 'Наименование', 'Код', 'Цена', 'Кол-во', 'Тип цены', 'Поступил в склад', 'Последнее поступление'
    ]
    worksheet.append(columns)

    # Заполнение данных
    products = Product.objects.all()
    for product in products:
        row = [
            product.id,
            product.category.name if product.category else 'N/A',
            product.name,
            product.articul,
            product.price,
            product.quantity,
            product.type_of_price.name if product.type_of_price else 'N/A',
            product.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            product.last_updated.strftime('%Y-%m-%d %H:%M:%S'),
        ]
        worksheet.append(row)

    # Применение границ к ячейкам
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Автоматическая подстройка ширины колонок
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Получение буквы колонки
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # Сохранение в объект BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Создание HTTP-ответа с Excel-файлом
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=price_list.xlsx'

    return response

def import_products(request):
    if request.method == 'POST':
        xlsx_file = request.FILES.get('file')
        if not xlsx_file.name.endswith('.xlsx'):
            messages.error(request, 'Это не файл XLSX')
            return redirect('import_products')

        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active

        # Словарь для хранения данных о продуктах с одинаковыми артикулами
        updated_products = defaultdict(lambda: {'quantity': 0, 'price': None, 'type_of_price': None, 'category': None, 'name': None})

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Извлечение данных из строки
            name, articul, quantity, price, type_of_price, category_name = row

            # Проверка, что обязательные поля не None
            if not all([name, articul, price, quantity, type_of_price, category_name]):
                messages.warning(request, f'Пропущены обязательные данные в строке: {row}')
                continue  # Пропустить строку, если отсутствуют данные

            # Найдите или создайте категорию
            category, created = Category.objects.get_or_create(name=category_name)

            # Найдите или создайте тип цены
            type_of_price_obj, created = TypeOfPrice.objects.get_or_create(name=type_of_price)

            # Если артикул уже встречался, накопить количество
            updated_products[articul]['quantity'] += quantity
            updated_products[articul]['price'] = price  # Можно также обновлять цену
            updated_products[articul]['type_of_price'] = type_of_price_obj
            updated_products[articul]['category'] = category
            updated_products[articul]['name'] = name

        # Обработка накопленных товаров
        for articul, product_data in updated_products.items():
            product = Product.objects.filter(articul=articul).first()

            if product:
                # Если продукт существует, обновить количество и другие поля
                product.quantity += product_data['quantity']
                product.price = product_data['price']  # Обновить цену, если нужно
                product.type_of_price = product_data['type_of_price']
                product.category = product_data['category']
                product.save()
            else:
                # Если продукт не существует, создать новый
                Product.objects.create(
                    name=product_data['name'],
                    articul=articul,
                    price=product_data['price'],
                    quantity=product_data['quantity'],
                    type_of_price=product_data['type_of_price'],
                    category=product_data['category']
                )

        messages.success(request, 'Товары успешно импортированы')
        return redirect('import_products')

    return render(request, 'pages/import_products.html')
